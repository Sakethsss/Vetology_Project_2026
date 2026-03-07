import os
import json
import time
import google.generativeai as genai
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
from sklearn.metrics import confusion_matrix
from openai import OpenAI
from openpyxl import load_workbook

load_dotenv()

BATCH_SIZE = 5  # safe starting point

#gemini call
def init_gemini_model(model_name: str = "gemini-2.5-flash-lite"):
 
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "GEMINI_API_KEY not found."
        )

    genai.configure(api_key=api_key)
    return genai.GenerativeModel(model_name)

#readfilepath CLi
def get_excel_file_path(prompt: str = "Enter path to Excel file: ") -> Path:

    while True:
        user_input = input(prompt).strip().strip('"').strip("'")
        path = Path(user_input)

        if not path.exists():
            print("File does not exist. Try again.")
            continue

        if path.suffix not in [".xlsx", ".xls"]:
            print("File must be an Excel file (.xlsx or .xls).")
            continue

        return path

#diseases extract
NON_DISEASE_COLS = {
    "caseid",
    "findings (original radiologist report)",
    "conclusions (original radiologist report)",
    "recommendations (original radiologist report)",
    "original radiologist",
}

def get_diseases_from_df(df):
    return [
        col for col in df.columns
        if col.lower() not in NON_DISEASE_COLS
    ]

#Prompt
def load_prompt_template():
    with open("prompt.json", "r", encoding="utf-8") as f:
        config = json.load(f)

    prompt = f"""
{config["role"]}

Task:
{config["task"]}

## Diseases to Evaluate:
{{disease_list}}

## Radiology Reports:
{{reports_text}}

Instructions:
{chr(10).join(config["instructions"])}

Output Requirements:
{chr(10).join(config["output_requirements"])}
"""

    return prompt

def build_batch_labeling_prompt(reports, diseases, prompt_template):
    disease_list = "\n".join(f"- {d}" for d in diseases)

    report_blocks = []
    for i, rpt in enumerate(reports, start=1):
        report_blocks.append(f"""
### Report {i}

Findings:
{rpt['findings'] or "No findings provided."}

Conclusion:
{rpt['conclusion'] or "No conclusion provided."}
""")

    reports_text = "\n".join(report_blocks)
    prompt = prompt_template.format(
    disease_list=disease_list,
    reports_text=reports_text
)
    return prompt


#gemini call
def call_gemini_batch(model, prompt):
    response = model.generate_content(
        prompt,
        generation_config={
            "temperature": 0.1,
            "max_output_tokens": 4096
        }
    )

    text = response.text.strip()

    if text.startswith("```"):
        text = text.replace("```json", "").replace("```", "").strip()

    return json.loads(text)

#labelling fn
def label_reports_from_excel(
    excel_path: Path,
    model,
    prompt_template,
    batch_size=BATCH_SIZE
):
    df = pd.read_excel(excel_path)
    diseases = get_diseases_from_df(df)

    for disease in diseases:
     if disease in df.columns:
        df[disease] = df[disease].astype("object")

    print(f"\nLabeling {len(df)} reports in batches of {batch_size}...\n")

    for start in range(0, len(df), batch_size):
        end = min(start + batch_size, len(df))
        batch_df = df.iloc[start:end]

        reports = []
        for _, row in batch_df.iterrows():
            findings = "" if pd.isna(row["Findings (original radiologist report)"]) else str(row["Findings (original radiologist report)"])
            conclusion = "" if pd.isna(row["Conclusions (original radiologist report)"]) else str(row["Conclusions (original radiologist report)"])

            reports.append({
                "findings": findings.strip(),
                "conclusion": conclusion.strip(),
            })

        prompt = build_batch_labeling_prompt(reports, diseases, prompt_template)

        try:
            batch_labels = call_gemini_batch(model, prompt)
        except Exception as e:
            print(f"⚠️ Batch failed ({start + 1}-{end}): {e}")
            continue

        # Fill labels back into original dataframe
        for i, labels in enumerate(batch_labels):
            for disease in diseases:
                df.at[start + i, disease] = labels.get(disease, "Normal")

        time.sleep(10)

    return df

#Confusion matrix
def generate_confusion_matrix(llm_file: Path, gold_file: Path, diseases):

    llm_df = pd.read_excel(llm_file)
    gold_df = pd.read_excel(gold_file)

    results = []

    for disease in diseases:

        y_pred = llm_df[disease].map(lambda x: 1 if str(x).strip().lower() == "abnormal" else 0)
        y_true = gold_df[disease].map(lambda x: 1 if str(x).strip().lower() == "abnormal" else 0)

        cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
        tn, fp, fn, tp = cm.ravel()

        results.append({
            "Condition": disease,
            "True Positive": tp,
            "False Negative": fn,
            "True Negative": tn,
            "False Positive": fp,
        })

    return pd.DataFrame(results)

#main fn
def main():
    print("Initializing Gemini...")
    model = init_gemini_model()
    print(" LLM initialized")
    prompt_template = load_prompt_template()
    
    input_file = get_excel_file_path(
        "Enter path to INPUT Excel file: "
    )
    print(f" Using file: {input_file}")

    labeled_df = label_reports_from_excel(
        excel_path=input_file,
        model=model,
        prompt_template=prompt_template
    )

    output_path = input_file.parent / "llm_labels_output.xlsx"
    labeled_df.to_excel(output_path, index=False)

    print(f"\n LLM labeling complete")
    print(f" Output saved to: {output_path}")

      # Ask for gold standard file
    gold_file = get_excel_file_path(
        "Enter path to GOLD STANDARD Excel file: "
    )
    
    diseases = get_diseases_from_df(labeled_df)
    confusion_df = generate_confusion_matrix(
        llm_file=output_path,
        gold_file=gold_file,
        diseases=diseases
    )

    confusion_output = input_file.parent / "confusion_matrix_output.xlsx"
    confusion_df.to_excel(confusion_output, index=False)

    wb = load_workbook(confusion_output)
    ws = wb.active

    ws["F1"] = "Sensitivity"
    ws["G1"] = "Specificity"
    ws["H1"] = "Total"
    ws["I1"] = "Positive Ground Truth"
    ws["J1"] = "Negative Ground Truth"
    ws["K1"] = "Ground Truth Check"

    for row in range(2, ws.max_row + 1):

     # Sensitivity = TP / (TP + FN)
     ws[f"F{row}"] = f"=IFERROR(B{row}/(B{row}+C{row}),0)"

     # Specificity = TN / (TN + FP)
     ws[f"G{row}"] = f"=IFERROR(D{row}/(D{row}+E{row}),0)"

     # Total
     ws[f"H{row}"] = f"=SUM(B{row}:E{row})"

     # Positive Ground Truth
     ws[f"I{row}"] = f"=B{row}+C{row}"

     # Negative Ground Truth
     ws[f"J{row}"] = f"=D{row}+E{row}"

     # Ground Truth Check
     ws[f"K{row}"] = f"=I{row}+J{row}"

     # Format as percentage
     ws[f"F{row}"].number_format = '0.00%'
     ws[f"G{row}"].number_format = '0.00%'

    wb.save(confusion_output)

    print(f"\n Confusion matrix saved to: {confusion_output}")


if __name__ == "__main__":
    main()